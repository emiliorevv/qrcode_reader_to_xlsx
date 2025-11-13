# excel_utils.py
from openpyxl import Workbook, load_workbook
from pathlib import Path
from datetime import datetime

EXCEL_FILE = "empleados_qr_actualizado.xlsx"
HEADERS = ["Nombre", "Número de empleado", "Área", "timestamp"]

def _ensure_sheet(ws):
    # Si la hoja está vacía, pone encabezados
    if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
        ws.append(HEADERS)

def append_employee_row(campos: dict):
    """
    campos = {"Nombre": "...", "Número de empleado": "...", "Área": "..."}
    """
    path = Path(EXCEL_FILE)
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active

    _ensure_sheet(ws)

    row = [campos["Nombre"], campos["Número de empleado"], campos["Área"],
           datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
    ws.append(row)
    wb.save(path)
